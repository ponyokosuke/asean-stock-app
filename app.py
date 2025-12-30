import streamlit as st
import pandas as pd
import time
import io
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Import existing logic
import data_processor

# 1. Page config
st.set_page_config(page_title="ASEAN Stock Analyzer", layout="wide")

# 2. Password Authentication logic
def check_password():
    def password_entered():
        if st.session_state["password"] == st.secrets["APP_PASSWORD"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.text_input("Please enter the password to access this app:", type="password", on_change=password_entered, key="password")
        return False
    elif not st.session_state["password_correct"]:
        st.text_input("Please enter the password to access this app:", type="password", on_change=password_entered, key="password")
        st.error("😕 Password incorrect")
        return False
    else:
        return True

if not check_password():
    st.stop()

# --- Initialize session state for processed data ---
# これにより、ボタンを押した後に画面がリセットされてもデータが残ります
if "final_excel" not in st.session_state:
    st.session_state["final_excel"] = None
if "df_preview" not in st.session_state:
    st.session_state["df_preview"] = None
if "processing_log" not in st.session_state:
    st.session_state["processing_log"] = ""

# --- UI Setup ---
st.title("📊 ASEAN Stock Financial & AI Analysis Tool")
st.markdown("Upload a stock list (CSV) to analyze data via Yahoo Finance and Gemini AI.")

# Sidebar Settings
with st.sidebar:
    st.header("Settings")
    if "GEMINI_API_KEY" in st.secrets:
        os.environ["GEMINI_API_KEY"] = st.secrets["GEMINI_API_KEY"]
        from google import genai
        data_processor.client = genai.Client(api_key=os.environ["GEMINI_API_KEY"])
        st.success("API Key loaded ✅")
    else:
        api_key = st.text_input("Gemini API Key", type="password")
        if api_key:
            os.environ["GEMINI_API_KEY"] = api_key
            from google import genai
            data_processor.client = genai.Client(api_key=api_key)

# File Upload
uploaded_file = st.file_uploader("Upload Stock List (CSV)", type=["csv"])
use_sample = st.checkbox("Use default list (asean_list.csv)")

# --- Main Logic ---
if st.button("Start Analysis 🚀"):
    target_csv = uploaded_file if uploaded_file else ("asean_list.csv" if use_sample else None)
    
    if target_csv is None:
        st.error("Please upload a CSV file.")
    else:
        status_text = st.empty()
        progress_bar = st.progress(0)
        
        # 全体のエラーハンドリング
        try:
            # 1. Read CSV
            df_input = pd.read_csv(target_csv, header=None)
            codes = df_input[0].astype(str).tolist()
            total_codes = len(codes)
            
            st.info(f"Target Stocks: {total_codes}")
            
            # 2. Data Retrieval Loop
            all_results = []
            error_count = 0
            
            for i, code in enumerate(codes):
                code = code.strip()
                status_text.text(f"Processing ({i+1}/{total_codes}): {code}...")
                progress_bar.progress((i + 1) / (total_codes + 1))
                
                # ★修正点: 個別の銘柄でエラーが起きても、ループを止めないようにここにtryを入れる
                try:
                    raw_data = data_processor.get_stock_data(code)
                    if raw_data:
                        processed_data = data_processor.extract_data(code, raw_data)
                        all_results.append(processed_data)
                    else:
                        print(f"Skipped {code}: No data found")
                except Exception as e:
                    print(f"Error processing {code}: {e}")
                    error_count += 1
                
                time.sleep(0.5) 
            
            # 3. AI Analysis
            if all_results:
                status_text.text("🤖 Running AI Segment Analysis...")
                try:
                    all_results = data_processor.batch_analyze_segments(all_results)
                except Exception as e:
                    st.warning(f"AI Analysis failed (skipping): {e}")
            
            progress_bar.progress(1.0)
            status_text.text("✅ Processing Completed!")

            # 4. Create DataFrame and Format
            if all_results:
                df = pd.DataFrame(all_results)
                # 重複列の削除
                df = df.loc[:, ~df.columns.duplicated()]
                
                # Excel用にフォーマット
                df = data_processor.format_for_excel(df)
                
                # Logic for columns and naming
                if "Sector /Industry" in df.columns:
                    df = df.drop(columns=["Sector /Industry"])
                df["Ref"] = range(1, len(df) + 1)
                
                empty_cols = ["Taka's comments", "Remarks", "Visited (V) / Meeting Proposal (MP)", "Access", "Last Communications", "Category Classification/\nShareInvestor", "Incorporated\n (IN / Year)", "Category Classification/SGX", "Sector & Industry/ SGX"]
                for col in empty_cols: df[col] = ""
                df["Listed 'o' / Non Listed \"x\""] = "o"

                yesterday_str = (datetime.now() - timedelta(days=1)).strftime("%b %d")
                final_stock_price_col = f"Stock Price ({yesterday_str}, Closing)"
                final_rate_col = f"Exchange Rate (to SGD) ({yesterday_str}, Closing)"
                
                df = df.rename(columns={"Stock Price": final_stock_price_col, "Exchange Rate": final_rate_col})

                # Final Reordering
                target_order = ["Ref", "Name of Company", "Code", "Listed 'o' / Non Listed \"x\"", "Taka's comments", "Remarks", "Visited (V) / Meeting Proposal (MP)", "Website", "Major Shareholders", "Currency", final_rate_col, "FY", "REVENUE SGD('000)", "Segments", "PROFIT ('000)", "GROSS PROFIT ('000)", "OPERATING PROFIT ('000)", "NET PROFIT (Group) ('000)", "NET PROFIT (Shareholders) ('000)", "Minority Interest ('000)", "Shareholders' Equity ('000)", "Total Equity ('000)", "TOTAL ASSET ('000)", "Debt/Equity(%)", "Loan ('000)", "Loan/Equity (%)", final_stock_price_col, "Shares Outstanding ('000)", "Market Cap ('000)", "Summary of Business", "Chairman / CEO", "Address", "Contact No.", "Access", "Last Communications", "Number of Employee Current", "Category Classification/YahooFin", "Sector & Industry/YahooFin", "Category Classification/\nShareInvestor", "Incorporated\n (IN / Year)", "Category Classification/SGX", "Sector & Industry/ SGX"]
                
                for col in target_order:
                    if col not in df.columns: df[col] = ""
                if "Number of Employee" in df.columns:
                    df = df.rename(columns={"Number of Employee": "Number of Employee Current"})
                
                df = df.reindex(columns=target_order)

                # --- 5. Excel Generation (STYLING) ---
                output = io.BytesIO()
                # Create raw excel first
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Data')
                
                # Load back to style
                output.seek(0)
                wb = load_workbook(output)
                ws = wb['Data']
                
                # Styling Logic
                yellow_fill = PatternFill(start_color="fefe99", end_color="fefe99", fill_type="solid")
                bold_font = Font(bold=True)
                right_align = Alignment(horizontal='right')

                for cell in ws[1]:
                    cell.fill = yellow_fill
                    cell.font = bold_font
                    col_name, col_idx = str(cell.value), cell.column
                    
                    fmt = None
                    if "('000)" in col_name: fmt = '#,##0;(#,##0)'
                    elif "(%)" in col_name or "%" in col_name: fmt = '0.00%'
                    elif "Stock Price" in col_name: fmt = '#,##0.000'
                    elif "Exchange Rate" in col_name: fmt = '0.0000'

                    if fmt or col_name == "FY":
                        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                            for c in row:
                                c.alignment = right_align
                                if fmt: c.number_format = fmt

                # Final Save to session state
                final_output = io.BytesIO()
                wb.save(final_output)
                
                # バイナリデータをセッションステートに保存
                st.session_state["final_excel"] = final_output.getvalue()
                st.session_state["df_preview"] = df
                st.session_state["processing_log"] = f"Processed {len(all_results)} stocks. (Errors: {error_count})"
            else:
                st.error("No data could be retrieved. Please check the stock codes.")

        except Exception as e:
            st.error(f"Critical Error: {e}")

# --- 6. Persistent Download Area ---
# 解析が終わった後にここが表示され続けます（再実行しても消えません）
if st.session_state["final_excel"] is not None:
    st.divider()
    st.success(f"Analysis Complete! {st.session_state['processing_log']}")
    
    st.download_button(
        label="📥 Download Excel File",
        data=st.session_state["final_excel"],
        file_name=f"asean_financial_data_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.subheader("Data Preview")
    st.dataframe(st.session_state["df_preview"])
